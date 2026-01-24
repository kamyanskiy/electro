"""Admin-only API endpoints."""

from uuid import UUID
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from app.core.models.user import User
from app.core.services.activation import ActivationService
from app.core.services.reading_service import ReadingService
from app.adapters.rest_api.schemas.users import UserResponse
from app.adapters.rest_api.schemas.readings import (
    ReadingWithUserResponse,
    AdminReadingsListResponse
)
from app.adapters.rest_api.dependencies import get_current_admin_user
from app.container import Container
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()


@router.get(
    "/admin/users/pending",
    response_model=list[UserResponse],
    summary="Get pending user activations",
    description="Get list of all users waiting for activation (admin only)."
)
async def get_pending_users(
    current_admin: User = Depends(get_current_admin_user),
    activation_service: ActivationService = Depends(lambda: Container.activation_service())
):
    """Get all inactive users waiting for activation."""
    inactive_users = await activation_service.get_pending_users()

    return [
        UserResponse(
            id=user.id,
            plot_number=user.plot_number,
            username=user.username,
            email=user.email,
            role=user.role.value,
            is_active=user.is_active,
            created_at=user.created_at,
            activated_at=user.activated_at
        )
        for user in inactive_users
    ]


@router.post(
    "/admin/users/{user_id}/activate",
    response_model=UserResponse,
    summary="Activate user",
    description="Activate a user account (admin only)."
)
async def activate_user(
    user_id: UUID,
    current_admin: User = Depends(get_current_admin_user),
    activation_service: ActivationService = Depends(lambda: Container.activation_service())
):
    """Activate a user account."""
    try:
        activated_user = await activation_service.activate_user(
            user_id=user_id,
            admin_id=current_admin.id
        )

        return UserResponse(
            id=activated_user.id,
            plot_number=activated_user.plot_number,
            username=activated_user.username,
            email=activated_user.email,
            role=activated_user.role.value,
            is_active=activated_user.is_active,
            created_at=activated_user.created_at,
            activated_at=activated_user.activated_at
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except PermissionError as e:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=str(e)
        )


@router.get(
    "/admin/readings",
    response_model=AdminReadingsListResponse,
    summary="Get all readings",
    description="Get all meter readings with optional filtering by month (admin only)."
)
async def get_all_readings(
    year: int | None = Query(default=None, description="Year to filter readings"),
    month: int | None = Query(default=None, ge=1, le=12, description="Month to filter readings (1-12)"),
    current_admin: User = Depends(get_current_admin_user),
    reading_service: ReadingService = Depends(lambda: Container.reading_service())
):
    """Get all readings with optional month filtering."""
    # If both year and month are provided, filter by month
    if year is not None and month is not None:
        readings_data = await reading_service.get_all_readings_by_month(year, month)
    # If only one is provided, return error
    elif year is not None or month is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Both year and month must be provided for filtering, or neither"
        )
    # Otherwise, get all readings
    else:
        readings_data = await reading_service.get_all_readings()

    reading_responses = [
        ReadingWithUserResponse(
            id=reading.id,
            user_id=reading.user_id,
            plot_number=plot_number,
            username=username,
            day_reading=reading.day_reading,
            night_reading=reading.night_reading,
            reading_date=reading.reading_date
        )
        for reading, plot_number, username in readings_data
    ]

    return AdminReadingsListResponse(
        readings=reading_responses,
        total=len(reading_responses)
    )


@router.get(
    "/admin/readings/export",
    summary="Export readings to Excel",
    description="Export meter readings to Excel file with optional filtering by month (admin only)."
)
async def export_readings(
    year: int | None = Query(default=None, description="Year to filter readings"),
    month: int | None = Query(default=None, ge=1, le=12, description="Month to filter readings (1-12)"),
    current_admin: User = Depends(get_current_admin_user),
    reading_service: ReadingService = Depends(lambda: Container.reading_service())
):
    """Export readings to Excel file."""
    # Get readings with same filtering logic as get_all_readings
    if year is not None and month is not None:
        readings_data = await reading_service.get_all_readings_by_month(year, month)
        filename = f"readings_{year}_{month:02d}.xlsx"
    elif year is not None or month is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Both year and month must be provided for filtering, or neither"
        )
    else:
        readings_data = await reading_service.get_all_readings()
        filename = f"readings_all_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Показания"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    headers = ["№", "Участок", "Пользователь", "Дата", "День (кВт⋅ч)", "Ночь (кВт⋅ч)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data
    for row_num, (reading, plot_number, username) in enumerate(readings_data, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=plot_number)
        ws.cell(row=row_num, column=3, value=username)
        ws.cell(row=row_num, column=4, value=reading.reading_date.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=5, value=float(reading.day_reading))
        ws.cell(row=row_num, column=6, value=float(reading.night_reading))

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
